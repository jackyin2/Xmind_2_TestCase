#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Software: PyCharm
# @Author  : Jack
# @FileName: excel.py
# @Time    : 2019/12/2 17:49

'''excel 操作'''


import openpyxl
from openpyxl.styles import Font
from tools.xmind_parse import l
from openpyxl.styles.colors import RED, BLUE, GREEN

class ExcelWriter():
    def __init__(self, excel_name, sheet_name="测试用例"):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = sheet_name
        self.excel_name = excel_name

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.excel_name)

    # 初始化标题
    def init_title(self):
        self.sheet['A1']= '测试编号'
        self.sheet['B1'] = '模块'
        self.sheet['C1'] = '功能'
        self.sheet['D1'] = '优先级'
        self.sheet['E1'] = '测试标题'
        self.sheet['F1'] = '测试步骤'
        self.sheet['G1'] = '期望结果'
        self.sheet['H1'] = '测试结果'
        self.sheet['I1'] = '备注'
        # 给用例标题加粗
        for col in range(1, 10):
            self.sheet.cell(row=1, column=col).font = Font(bold=True)

    def write_rows(self):
        global l
        row = 2
        # 按行赋值
        for i in l:
            self.sheet.cell(row=row, column=1).value = row-1
            self.sheet.cell(row=row, column=2).value = i.model
            self.sheet.cell(row=row, column=3).value = i.function
            self.sheet.cell(row=row, column=4).value = i.priority
            self.sheet.cell(row=row, column=5).value = i.title
            self.sheet.cell(row=row, column=6).value = i.step
            self.sheet.cell(row=row, column=7).value = i.expect
            self.sheet.cell(row=row, column=8).value = i.result
            self.sheet.cell(row=row, column=9).value = i.notes
            if i.result == 'Pass':
                self.sheet.cell(row=row, column=8).font = Font(color=GREEN)
            elif i.result == 'Fail':
                self.sheet.cell(row=row, column=8).font = Font(color=RED)
            row += 1

    def show_report(self):
        global l
        success = norun = fail = 0
        for i in l:
            if i.result == 'Pass':
                success += 1
            elif i.result == 'Fail':
                fail += 1
            else:
                norun += 1
        print('总计成功导出用例：{}个， 成功：{}个， 失败：{}个， 未执行：{}个'.format(len(l), success, fail, norun))


if __name__ == "__main__":
    with ExcelWriter('xxx.xlsx') as ew:
        ew.init_title()
        ew.write_rows()